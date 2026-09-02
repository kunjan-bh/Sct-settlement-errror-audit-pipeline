import os
import re
from collections import defaultdict
from io import BytesIO
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

from app.extensions import db
from app.models.batch import Batch
from app.models.transaction import Transaction
from app.models.issue_status import IssueStatus
from app.services.dashboard_service import build_dashboard
from app.services.excel_ingest import _find_header_index
from app.services.status_utils import normalize_txn_status
from app.services.error_classification import (
    ENTITY_TYPE_LABELS,
    SIDE_LABELS,
    build_error_classification,
)
from app.services.settlement_type_service import build_settlement_type_report, build_settlement_type_mid_rows
from app.services.adhoc_settlement_service import build_adhoc_settlement_type_full
from app.services.transaction_reconcile import (
    TXN_REPORT_HEADERS,
    TXN_REPORT_KEYS,
    attach_transactions,
    build_transaction_index,
)


_DATA_STYLE_NAME = "sct_data"
_SOLVED_STYLE_NAME = "sct_solved"
_PENDING_STYLE_NAME = "sct_pending"
_NEUTRAL_STATUS_STYLE_NAME = "sct_status_blank"


def _register_data_styles(wb, font_regular, border_cell, align_left, align_center,
                          fill_solved, font_solved, fill_pending, font_pending):
    """
    Register the Processed Records cell styles on the workbook once.

    openpyxl resolves `cell.style = "name"` through the workbook's named-style
    table instead of hashing a fresh style object per assignment, which is the
    difference between ~2 minutes and ~25 seconds on a 14k-row batch. The
    styles are identical to the ones they replace, so the sheet looks the same.

    Idempotent: openpyxl raises if a name is added twice, and a workbook is
    built per report, but guarding keeps this safe if that ever changes.
    """
    def add(name, font, fill=None, alignment=None):
        if name in wb.named_styles:
            return
        style = NamedStyle(name=name)
        style.font = font
        style.border = border_cell
        style.alignment = alignment or align_left
        if fill is not None:
            style.fill = fill
        wb.add_named_style(style)

    add(_DATA_STYLE_NAME, font_regular)
    add(_SOLVED_STYLE_NAME, font_solved, fill_solved, align_center)
    add(_PENDING_STYLE_NAME, font_pending, fill_pending, align_center)
    add(_NEUTRAL_STATUS_STYLE_NAME, font_regular, None, align_center)


def _is_phantom_issue(issue_obj, override_obj, txn_original_status) -> bool:
    """
    True for a row that was never an issue in the first place: a settlement
    that succeeded, with no IssueStatus behind it and no per-MID override.

    The dashboard deliberately does not create an IssueStatus for a success
    row (services/dashboard_service.py: `if not status_row and
    txn_status_name != "success"`), so no issue card is ever rendered for
    one. Ops therefore cannot solve or exclude it -- there is nothing to
    click. The report used to fall back to "pending" for exactly these rows,
    which turned every successful settlement into an outstanding issue: on a
    real batch that was 12,566 phantom "Unclassified / Pending" rows burying
    the ~30 real ones, and it made a fully worked batch look untouched.

    A success row that ops DID flag deliberately (it has an IssueStatus row,
    or a MID override) is not phantom and still reports its real status.
    """
    return issue_obj is None and override_obj is None and txn_original_status == "success"


# batch_id -> (signature, bytes). One entry per batch, replaced whenever the
# signature moves, so this cannot grow without bound in any realistic use.
_REPORT_CACHE: dict[int, tuple[tuple, bytes]] = {}


def _report_signature(batch_id: int) -> tuple:
    """
    Everything that can change this batch's report, cheap enough to compute on
    every request.

    Deliberately not a timestamp on the batch: ops edits issue statuses and
    per-MID overrides long after the batch row itself last changed, and each of
    those alters the Solved Status column and the Issue Summary counts. Keyed
    on the issue table's own high-water mark instead, plus the row counts that
    reveal an ingest or a retry-reconciliation pass.
    """
    batch = Batch.query.get(batch_id)
    if batch is None:
        return ()

    txn_count, retry_count = (
        db.session.query(
            db.func.count(Transaction.id),
            db.func.sum(db.case((Transaction.retry_resolved.is_(True), 1), else_=0)),
        )
        .filter(Transaction.batch_id == batch_id)
        .one()
    )
    issue_count, issue_high_water = (
        db.session.query(db.func.count(IssueStatus.id), db.func.max(IssueStatus.updated_at))
        .filter(IssueStatus.batch_id == batch_id)
        .one()
    )
    return (
        batch.status, batch.finished_at, batch.name,
        txn_count, retry_count or 0,
        issue_count, issue_high_water,
    )


def invalidate_report_cache(batch_id: int | None = None) -> None:
    """Drop a cached report (or all of them). Not currently called -- the
    signature check already catches every change we know of -- but here so a
    future writer has an obvious lever instead of reaching into the dict."""
    if batch_id is None:
        _REPORT_CACHE.clear()
    else:
        _REPORT_CACHE.pop(batch_id, None)


def generate_report_bytes(batch_id: int) -> bytes:
    """
    Cached front door to _build_report_bytes.

    Building this workbook takes the better part of a minute on a 14k-row
    batch -- openpyxl serialising half a million cells -- and the summary
    email attaches the same bytes the Download button produces. Without the
    cache, pressing Send rebuilt from scratch and looked like it had hung.

    The cache is keyed on a signature of the underlying data, so an edit to an
    issue status invalidates it immediately; it never serves a stale report.
    In-process only, so a restart simply rebuilds.
    """
    signature = _report_signature(batch_id)
    cached = _REPORT_CACHE.get(batch_id)
    if cached is not None and cached[0] == signature and signature:
        return cached[1]

    data = _build_report_bytes(batch_id)
    if signature:
        _REPORT_CACHE[batch_id] = (signature, data)
    return data


def _build_report_bytes(batch_id: int) -> bytes:
    """
    Generates the 5-sheet Excel report workbook for the given batch_id and
    returns its raw bytes.
    Sheet 1: "Overview"
    Sheet 2: "Issue Summary"
    Sheet 3: "Lo Status"
    Sheet 4: "Processed Records" (second-to-last -- the raw row dump, kept
             behind the summary sheets so those open first)
    Sheet 5: "Error Classify"
    """
    batch = Batch.query.get_or_404(batch_id)
    transactions = Transaction.query.filter_by(batch_id=batch_id).all()
    issues = IssueStatus.query.filter_by(batch_id=batch_id).all()
    dashboard = build_dashboard(batch_id)

    # Fast lookup for IssueStatus status & comment:
    # issue_key = (side, partner_name, category, txn_status)
    issue_status_map = {
        (i.side, i.partner_name, i.category, i.txn_status): i
        for i in issues
    }

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active

    # --- Styles ---
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

    fill_solved = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_solved = Font(name="Calibri", size=10, color="006100", bold=True)

    fill_pending = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    font_pending = Font(name="Calibri", size=10, color="9C5700", bold=True)

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # ==========================================
    # SHEET 1: Overview
    # ==========================================
    ws1 = wb.create_sheet(title="Overview")
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "SmartQR Settlement Processing Report — Overview"
    ws1["A1"].font = font_title

    # Metadata Block
    ws1["A3"] = "Batch Name:"
    ws1["B3"] = batch.name
    ws1["A4"] = "Created Date:"
    ws1["B4"] = batch.created_at.strftime("%Y-%m-%d %H:%M:%S") if batch.created_at else ""
    ws1["A5"] = "Status:"
    ws1["B5"] = batch.status.upper()
    ws1["A6"] = "Total Transactions:"
    ws1["B6"] = dashboard["totals"]["total_transactions"]
    ws1["A7"] = "Pending Transactions:"
    ws1["B7"] = dashboard["totals"]["pending"]
    ws1["A8"] = "Lo Progress Transactions:"
    ws1["B8"] = dashboard["totals"]["lo_progress"]
    ws1["A9"] = "Settlement Failed:"
    ws1["B9"] = dashboard["totals"]["settlement_failed"]
    ws1["A10"] = "Transaction Failed (SCT):"
    ws1["B10"] = dashboard["totals"]["transaction_failed"]
    ws1["A11"] = "Reprocessed & Settled (excluded):"
    ws1["B11"] = dashboard["totals"].get("retry_resolved", 0)
    ws1["A12"] = "Batch Notes:"
    ws1["B12"] = batch.notes or ""

    for row_idx in range(3, 13):
        ws1[f"A{row_idx}"].font = font_bold
        ws1[f"B{row_idx}"].font = font_regular

    # Table 1: Partner Summary (aggregators first, then banks/wallets).
    # Counts come from the dashboard's three per-status issue lists -- there
    # is no precomputed "issue_count", the unique-issue total is their length.
    ws1["A15"] = "Partner Summary"
    ws1["A15"].font = font_section

    headers_agg = ["Partner", "Type", "Failed", "Pending", "Lo Progress", "Unique Issues"]
    for col_idx, h in enumerate(headers_agg, start=1):
        cell = ws1.cell(row=16, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    curr_row = 17
    partner_rows = (
        [("Aggregator", p) for p in dashboard["aggregator_summary"]]
        + [("Bank / Wallet", p) for p in dashboard["bank_summary"]]
    )
    for type_label, agg in partner_rows:
        unique_issues = (
            len(agg["issues_failed"]) + len(agg["issues_pending"]) + len(agg["issues_lo_progress"])
        )
        ws1.cell(row=curr_row, column=1, value=agg["partner_name"]).alignment = align_left
        ws1.cell(row=curr_row, column=2, value=type_label).alignment = align_left
        ws1.cell(row=curr_row, column=3, value=agg["failed"]).alignment = align_right
        ws1.cell(row=curr_row, column=4, value=agg["pending"]).alignment = align_right
        ws1.cell(row=curr_row, column=5, value=agg["lo_progress"]).alignment = align_right
        ws1.cell(row=curr_row, column=6, value=unique_issues).alignment = align_right
        for col_idx in range(1, 7):
            ws1.cell(row=curr_row, column=col_idx).font = font_regular
            ws1.cell(row=curr_row, column=col_idx).border = border_cell
        curr_row += 1

    # Table 2: SCT Summary
    curr_row += 2
    ws1.cell(row=curr_row, column=1, value="SCT Summary").font = font_section
    curr_row += 1

    headers_sct = ["SCT Issue Category", "Txn Status", "Count"]
    for col_idx, h in enumerate(headers_sct, start=1):
        cell = ws1.cell(row=curr_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    curr_row += 1

    sct = dashboard["sct_summary"]
    sct_issues = sct["issues_failed"] + sct["issues_pending"] + sct["issues_lo_progress"]
    for sct_issue in sct_issues:
        ws1.cell(row=curr_row, column=1, value=sct_issue["category"]).alignment = align_left
        ws1.cell(
            row=curr_row, column=2,
            value=sct_issue["txn_status"].replace("_", " ").title(),
        ).alignment = align_center
        ws1.cell(row=curr_row, column=3, value=sct_issue["count"]).alignment = align_right
        for col_idx in range(1, 4):
            ws1.cell(row=curr_row, column=col_idx).font = font_regular
            ws1.cell(row=curr_row, column=col_idx).border = border_cell
        curr_row += 1

    # ==========================================
    # SHEET 2: Issue Summary
    # ==========================================
    ws3 = wb.create_sheet(title="Issue Summary")
    ws3.views.sheetView[0].showGridLines = True

    ws3["A1"] = "Issue Summary — Breakdown by Partner & SCT"
    ws3["A1"].font = font_title

    headers_summary = ["Entity / Partner", "Issue Category", "Affected Txns", "Solved Status", "Ops Remarks"]
    for col_idx, h in enumerate(headers_summary, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    curr_row = 4
    
    from collections import defaultdict
    summary_counts = defaultdict(int)
    summary_comments = {}
    # Sheet 4 accumulator: a row belongs on Lo Status if the transaction
    # itself came in as "In progress", OR if ops manually moved the issue to
    # Lo Progress. Keying it off ops status alone (as this used to) left the
    # sheet empty for every batch nobody had hand-flagged.
    lo_counts = defaultdict(int)
    lo_comments = {}

    for txn in transactions:
        if txn.retry_resolved:
            continue  # settled by a later reprocess; not an outstanding issue

        side = txn.error_side
        partner = txn.partner_name if side != "sct" else "SCT"
        category = txn.error_category or "Unclassified"
        txn_original_status = normalize_txn_status(txn.status)

        issue_obj = issue_status_map.get((side, txn.partner_name if side != "sct" else None, category, txn_original_status))
        
        override_obj = None
        if issue_obj and issue_obj.mid_overrides and txn.mid in issue_obj.mid_overrides:
            raw = issue_obj.mid_overrides[txn.mid]
            if isinstance(raw, dict):
                override_obj = raw
            else:
                override_obj = {"status": raw, "remark": ""}

        # An Issue Summary lists issues. A settlement that succeeded and was
        # never raised as one does not belong here at all.
        if _is_phantom_issue(issue_obj, override_obj, txn_original_status):
            continue

        eff_status = (override_obj["status"] if override_obj else None) or (issue_obj.status if issue_obj else "pending")
        if eff_status == "exclude":
            continue
            
        key = (side, partner, category, eff_status)
        summary_counts[key] += 1
        if key not in summary_comments:
             summary_comments[key] = issue_obj.comment if issue_obj else ""

        if side != "sct" and (txn_original_status == "lo_progress" or eff_status == "lo_progress"):
            lo_key = (partner, category, txn_original_status, eff_status)
            lo_counts[lo_key] += 1
            if lo_key not in lo_comments:
                lo_comments[lo_key] = issue_obj.comment if issue_obj else ""

    # 1. Aggregator & Bank issues
    agg_keys = sorted([k for k in summary_counts.keys() if k[0] != "sct"])
    for key in agg_keys:
        side, partner, category, eff_status = key
        count = summary_counts[key]
        
        ws3.cell(row=curr_row, column=1, value=partner).alignment = align_left
        ws3.cell(row=curr_row, column=2, value=category).alignment = align_left
        ws3.cell(row=curr_row, column=3, value=count).alignment = align_right
        
        status_val = eff_status.replace("_", " ").title()
        status_cell = ws3.cell(row=curr_row, column=4, value=status_val)
        status_cell.alignment = align_center
        if status_val == "Solved":
            status_cell.fill = fill_solved
            status_cell.font = font_solved
        else:
            status_cell.fill = fill_pending
            status_cell.font = font_pending

        ws3.cell(row=curr_row, column=5, value=summary_comments[key] or "").alignment = align_left

        for c in range(1, 6):
            if c != 4:  # status cell is formatted separately
                ws3.cell(row=curr_row, column=c).font = font_regular
            ws3.cell(row=curr_row, column=c).border = border_cell
        curr_row += 1

    # 2. SCT issues
    sct_keys = sorted([k for k in summary_counts.keys() if k[0] == "sct"])
    for key in sct_keys:
        side, partner, category, eff_status = key
        count = summary_counts[key]
        
        ws3.cell(row=curr_row, column=1, value="SCT").alignment = align_left
        ws3.cell(row=curr_row, column=2, value=category).alignment = align_left
        ws3.cell(row=curr_row, column=3, value=count).alignment = align_right
        
        status_val = eff_status.replace("_", " ").title()
        status_cell = ws3.cell(row=curr_row, column=4, value=status_val)
        status_cell.alignment = align_center
        if status_val == "Solved":
            status_cell.fill = fill_solved
            status_cell.font = font_solved
        else:
            status_cell.fill = fill_pending
            status_cell.font = font_pending

        ws3.cell(row=curr_row, column=5, value=summary_comments[key] or "").alignment = align_left

        for c in range(1, 6):
            if c != 4:  # status cell is formatted separately
                ws3.cell(row=curr_row, column=c).font = font_regular
            ws3.cell(row=curr_row, column=c).border = border_cell
        curr_row += 1

    # ==========================================
    # SHEET 3: Lo Status
    # ==========================================
    ws4 = wb.create_sheet(title="Lo Status")
    ws4.views.sheetView[0].showGridLines = True

    ws4["A1"] = "Lo Status Tracking"
    ws4["A1"].font = font_title

    headers_lo = [
        "Aggregator / Wallet", "Issue Category", "Txn Status",
        "Affected Txns", "Solved Status", "Ops Remarks",
    ]
    for col_idx, h in enumerate(headers_lo, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    curr_row_lo = 4
    for key in sorted(lo_counts.keys(), key=lambda k: (-lo_counts[k], k[0], k[1])):
        partner, category, txn_status, eff_status = key

        ws4.cell(row=curr_row_lo, column=1, value=partner).alignment = align_left
        ws4.cell(row=curr_row_lo, column=2, value=category).alignment = align_left
        ws4.cell(
            row=curr_row_lo, column=3, value=txn_status.replace("_", " ").title(),
        ).alignment = align_center
        ws4.cell(row=curr_row_lo, column=4, value=lo_counts[key]).alignment = align_right

        status_val = eff_status.replace("_", " ").title()
        status_cell = ws4.cell(row=curr_row_lo, column=5, value=status_val)
        status_cell.alignment = align_center
        if status_val == "Solved":
            status_cell.fill = fill_solved
            status_cell.font = font_solved
        else:
            status_cell.fill = fill_pending
            status_cell.font = font_pending

        ws4.cell(row=curr_row_lo, column=6, value=lo_comments[key] or "").alignment = align_left

        for c in range(1, 7):
            if c != 5:
                ws4.cell(row=curr_row_lo, column=c).font = font_regular
            ws4.cell(row=curr_row_lo, column=c).border = border_cell
        curr_row_lo += 1

    # ==========================================
    # SHEET 4: Processed Records (second-to-last, ahead of Error Classify)
    # ==========================================
    ws2 = wb.create_sheet(title="Processed Records")
    ws2.views.sheetView[0].showGridLines = True

    # Read original columns from file if available, or extract from transactions
    orig_headers = []
    if batch.input_file_path and os.path.exists(batch.input_file_path):
        try:
            h_idx = _find_header_index(batch.input_file_path)
            raw_df = pd.read_excel(batch.input_file_path, header=h_idx, nrows=1)
            orig_headers = [str(c).strip() for c in raw_df.columns if pd.notna(c)]
        except Exception:
            pass

    if not orig_headers:
        orig_headers = ["SN", "Transaction Date", "Acquirer Name", "MID", "Merchant Name", "Txn Amount", "Service Charge", "Settled By", "STAN", "CRRN", "CR Transaction ID", "Bank/Wallet Name", "Bank Account/Wallet ID", "Account Name", "Remarks 1", "Remarks 2", "Status Code", "Status"]

    # "Retry Settled" goes last so the existing column offsets (used for the
    # Solved Status formatting below) stay put.
    appended_headers = [
        "Aggregator", "Issue Category", "Solved Status", "Solved Remarks",
        "Batch", "Timestamp", "Retry Settled",
    ]
    full_headers = orig_headers + appended_headers

    # Registered once on the workbook; cells reference them by name below.
    solved_status_col = len(orig_headers) + 3
    _register_data_styles(wb, font_regular, border_cell, align_left, align_center,
                          fill_solved, font_solved, fill_pending, font_pending)

    for col_idx, h in enumerate(full_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    row_idx = 2
    for txn in transactions:
        extra = txn.extra_data or {}
        # Populate original column values
        row_vals = []
        for h in orig_headers:
            if h == "MID":
                row_vals.append(txn.mid)
            elif h == "Merchant Name":
                row_vals.append(txn.merchant_name)
            elif h == "Remarks 1":
                row_vals.append(txn.remark)
            elif h == "Status Code":
                row_vals.append(txn.status_code)
            elif h == "Status":
                row_vals.append(txn.status)
            else:
                row_vals.append(extra.get(h, ""))

        # Appended columns lookup
        txn_original_status = normalize_txn_status(txn.status)
        issue_obj = issue_status_map.get((txn.error_side, txn.partner_name if txn.error_side != "sct" else None, txn.error_category, txn_original_status))

        override_obj = None
        if issue_obj and issue_obj.mid_overrides and txn.mid in issue_obj.mid_overrides:
            raw = issue_obj.mid_overrides[txn.mid]
            # Support both old string format and new {status, remark} format
            if isinstance(raw, dict):
                override_obj = raw
            else:
                override_obj = {"status": raw, "remark": ""}

        eff_status = (override_obj["status"] if override_obj else None) or (issue_obj.status if issue_obj else "pending")

        # Skip transactions whose issue (or override) is marked as 'exclude'
        if eff_status == "exclude":
            continue

        # Processed Records is the full preserved copy of the input, so the
        # row stays -- but a settlement nobody was ever asked to work on has
        # no ops status to report. Blank, not "Pending".
        if _is_phantom_issue(issue_obj, override_obj, txn_original_status):
            solved_status = ""
            solved_remarks = ""
        else:
            solved_status = eff_status.replace("_", " ").title()
            # Override remark takes priority over group comment
            if override_obj and override_obj.get("remark"):
                solved_remarks = f"[MID Override] {override_obj['remark']}"
            else:
                solved_remarks = issue_obj.comment if issue_obj else ""

        row_vals.extend([
            txn.partner_name or "N/A",
            txn.error_category or "Unclassified",
            solved_status,
            solved_remarks,
            batch.name,
            batch.created_at.strftime("%Y-%m-%d %H:%M:%S") if batch.created_at else "",
            # The original row is preserved either way -- this column is how a
            # reader tells "still failed" from "failed, then reprocessed OK".
            "Yes" if txn.retry_resolved else "",
        ])

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            # One named-style assignment rather than setting font, border and
            # alignment separately. Assigning a style object makes openpyxl
            # hash the whole thing and look it up in the workbook's style
            # table; on this sheet that was ~1.7M assignments and about two
            # minutes, which is what made sending the summary email feel
            # like it had hung. Same appearance, ~5x faster.
            cell.style = _DATA_STYLE_NAME

            # Solved Status is the one column worth colouring, and there are
            # only a few thousand of them, so the cost here is irrelevant.
            if col_idx == solved_status_col:
                if val == "Solved":
                    cell.style = _SOLVED_STYLE_NAME
                elif val:
                    cell.style = _PENDING_STYLE_NAME
                else:
                    # Never an issue: centred like its neighbours, but
                    # unshaded -- amber would read as outstanding work.
                    cell.style = _NEUTRAL_STATUS_STYLE_NAME
        row_idx += 1

    # ==========================================
    # SHEET 5: Error Classify (last sheet)
    # Chart view over the same data the dashboard's Error Classification tab
    # and the standalone extract use (build_error_classification) -- so the
    # numbers behind the pictures can't drift from the raw extract. The raw
    # per-category table stays in the standalone extract (write_error_classify_
    # sheet) where it's meant to be filtered/pivoted; here it's visuals only.
    # ==========================================
    ws5 = wb.create_sheet(title="Error Classify")
    write_error_classify_charts_sheet(ws5, build_error_classification(batch_id))

    # Auto-adjust column widths across all sheets
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if val_str:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==========================================
# Settlement Type Report (date range + ad-hoc Document Analysis)
# ==========================================

_METHOD_LABEL_ORDER = [
    ("real_time", "Real Time"),
    ("system_default", "System Default"),
    ("on_call", "On Call"),
    ("unknown", "Unknown"),
]


def _write_settlement_type_workbook(
    meta: list[tuple[str, object]],
    data: dict,
    mid_rows: list[dict],
    reconciliation: dict | None = None,
) -> bytes:
    """
    Shared layout for both the date-range Settlement Type report and the
    one-file Document Analysis report -- same shape either way:

      Sheet 1 "Summary": title + metadata block, then two tables --
        Settlement Method Split by Count and by Amount.
      Sheet 2 "Entity Breakdown": one row per entity (aggregator/bank/
        wallet/SCT), same columns as the on-page table.
      Sheet 3+: one sheet per entity, listing every MID that settled under
        it -- see _write_entity_mid_sheets. This detail is report-only, by
        request: neither on-page view shows it.

    `meta` carries whatever's specific to the caller (date range + batches
    included, or file name + rows read) as (label, value) pairs.

    `reconciliation` is the tally from transaction_reconcile.attach_
    transactions, present only when a Transaction List was uploaded
    alongside. When it is, Summary gains a trace-rate block and the entity
    sheets gain their side-by-side transaction table; when it isn't, the
    workbook is byte-for-byte the report it was before.
    """
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    kpis = data["kpis"]
    method_breakdown = data["method_breakdown"]
    method_amount_breakdown = data["method_amount_breakdown"]
    entities = data["entities"]

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    # ---- Sheet 1: Summary ----
    ws1 = wb.create_sheet(title="Summary")
    ws1["A1"] = "Settlement Type Report — Summary"
    ws1["A1"].font = font_title

    row = 3
    for label, value in meta:
        ws1.cell(row=row, column=1, value=label).font = font_bold
        ws1.cell(row=row, column=2, value=value).font = font_regular
        row += 1

    total_settled = kpis["total_settled"]
    total_amount = kpis["total_amount_settled"]

    # Table A: by count
    row += 1
    ws1.cell(row=row, column=1, value="Settlement Method Split — By Count").font = font_section
    row += 1
    for col_idx, h in enumerate(["Method", "Count", "% of Total"], start=1):
        cell = ws1.cell(row=row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    row += 1
    for key, label in _METHOD_LABEL_ORDER:
        count = method_breakdown.get(key, 0)
        ws1.cell(row=row, column=1, value=label).alignment = align_left
        ws1.cell(row=row, column=2, value=count).alignment = align_right
        pct_cell = ws1.cell(row=row, column=3, value=(count / total_settled) if total_settled else 0)
        pct_cell.alignment = align_right
        pct_cell.number_format = "0.0%"
        for c in (1, 2, 3):
            ws1.cell(row=row, column=c).font = font_regular
            ws1.cell(row=row, column=c).border = border_cell
        row += 1
    ws1.cell(row=row, column=1, value="Total").font = font_bold
    ws1.cell(row=row, column=2, value=total_settled).font = font_bold
    ws1.cell(row=row, column=3, value=1.0 if total_settled else 0).font = font_bold
    ws1.cell(row=row, column=3).number_format = "0.0%"
    for c in (1, 2, 3):
        ws1.cell(row=row, column=c).border = border_cell
        ws1.cell(row=row, column=c).alignment = align_right if c > 1 else align_left
    row += 1

    # Table B: by amount
    row += 2
    ws1.cell(row=row, column=1, value="Settlement Method Split — By Amount").font = font_section
    row += 1
    for col_idx, h in enumerate(["Method", "Amount", "% of Total"], start=1):
        cell = ws1.cell(row=row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    row += 1
    for key, label in _METHOD_LABEL_ORDER:
        amount = method_amount_breakdown.get(key, 0.0)
        ws1.cell(row=row, column=1, value=label).alignment = align_left
        amt_cell = ws1.cell(row=row, column=2, value=amount)
        amt_cell.alignment = align_right
        amt_cell.number_format = "#,##0.00"
        pct_cell = ws1.cell(row=row, column=3, value=(amount / total_amount) if total_amount else 0)
        pct_cell.alignment = align_right
        pct_cell.number_format = "0.0%"
        for c in (1, 2, 3):
            ws1.cell(row=row, column=c).font = font_regular
            ws1.cell(row=row, column=c).border = border_cell
        row += 1
    ws1.cell(row=row, column=1, value="Total").font = font_bold
    total_amt_cell = ws1.cell(row=row, column=2, value=total_amount)
    total_amt_cell.font = font_bold
    total_amt_cell.number_format = "#,##0.00"
    ws1.cell(row=row, column=3, value=1.0 if total_amount else 0).font = font_bold
    ws1.cell(row=row, column=3).number_format = "0.0%"
    for c in (1, 2, 3):
        ws1.cell(row=row, column=c).border = border_cell
        ws1.cell(row=row, column=c).alignment = align_right if c > 1 else align_left

    # Table C: transaction trace rate -- only when a Transaction List was
    # uploaded. An unmatched settlement is not necessarily an error (the two
    # exports can cover different windows), so this is reported as a rate to
    # eyeball, not as a pass/fail.
    if reconciliation:
        row += 3
        ws1.cell(row=row, column=1, value="Transaction Reconciliation").font = font_section
        row += 1
        traced = reconciliation["matched"]
        settled_rows = reconciliation["settled_rows"]
        recon_lines = [
            ("Transaction File:", reconciliation["file_name"]),
            ("Transactions Read:", reconciliation["txn_rows"]),
            ("Transaction File Covers:", reconciliation.get("txn_window") or "—"),
            ("Settlement File Covers:", data.get("window") or "—"),
            ("Settlements Traced:", traced),
            ("Settlements Not Traced:", reconciliation["unmatched"]),
        ]
        by_network = reconciliation.get("by_network") or {}
        if by_network:
            recon_lines.append((
                "Transactions by Network:",
                ", ".join(f"{net} {count:,}" for net, count in sorted(by_network.items())),
            ))
        for label, value in recon_lines:
            ws1.cell(row=row, column=1, value=label).font = font_bold
            ws1.cell(row=row, column=2, value=value).font = font_regular
            row += 1
        ws1.cell(row=row, column=1, value="Trace Rate:").font = font_bold
        rate_cell = ws1.cell(row=row, column=2, value=(traced / settled_rows) if settled_rows else 0)
        rate_cell.font = font_bold
        rate_cell.number_format = "0.0%"

        if by_network.get("NQR"):
            row += 2
            nqr_note = ws1.cell(
                row=row, column=1,
                value=(
                    f"Note: {by_network['NQR']:,} of these transactions are NQR-issued. Their "
                    "transaction id and CRRN are bank references (e.g. PRVUNPKA-1219580) that the "
                    "settlement file does not record anywhere, so they cannot be traced by key and "
                    "are left blank. Only SmartQR transactions, whose id is the settlement file's "
                    "STAN and CRRN joined, can be traced."
                ),
            )
            nqr_note.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
            nqr_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws1.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=3)
            row += 2

        if not traced:
            # Silence here is what made this look like the upload was ignored.
            row += 2
            note = ws1.cell(
                row=row, column=1,
                value=(
                    "No settlement could be traced to a transaction. The two files most likely "
                    "cover different windows — compare the two 'Covers' rows above. A settlement "
                    "run spans a cycle (e.g. 24 Aug 10:00 AM to 25 Aug 09:59 AM), so a "
                    "transaction export for a single calendar day will only partly overlap it."
                ),
            )
            note.font = Font(name="Calibri", size=9, italic=True, color="B45309")
            note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws1.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=3)

    for col_idx, width in enumerate([32, 30, 14], start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ---- Sheet 2: Entity Breakdown ----
    ws2 = wb.create_sheet(title="Entity Breakdown")
    ws2["A1"] = "Settlement Type Report — Entity Breakdown"
    ws2["A1"].font = font_title
    ws2["A2"] = (
        "Of everything that settled successfully, how it settled — by aggregator, "
        "bank/wallet, and SCT."
    )
    ws2["A2"].font = Font(name="Calibri", size=9, italic=True, color="808080")

    headers_entity = [
        "Entity", "Type", "Real Time", "System Default", "On Call", "Unknown", "Total",
        "Total Amount",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers_entity, start=1):
        cell = ws2.cell(row=header_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    row = header_row + 1
    for ent in entities:
        ws2.cell(row=row, column=1, value=ent["entity"]).alignment = align_left
        ws2.cell(
            row=row, column=2, value=ENTITY_TYPE_LABELS.get(ent["entity_type"], ent["entity_type"]),
        ).alignment = align_left
        ws2.cell(row=row, column=3, value=ent["real_time"]).alignment = align_right
        ws2.cell(row=row, column=4, value=ent["system_default"]).alignment = align_right
        ws2.cell(row=row, column=5, value=ent["on_call"]).alignment = align_right
        ws2.cell(row=row, column=6, value=ent["unknown"]).alignment = align_right
        ws2.cell(row=row, column=7, value=ent["total"]).alignment = align_right
        amt_cell = ws2.cell(row=row, column=8, value=ent.get("amount", 0.0))
        amt_cell.alignment = align_right
        amt_cell.number_format = "#,##0.00"
        for c in range(1, 9):
            ws2.cell(row=row, column=c).font = font_regular
            ws2.cell(row=row, column=c).border = border_cell
        row += 1

    ws2.freeze_panes = ws2.cell(row=header_row + 1, column=1)
    if entities:
        ws2.auto_filter.ref = f"A{header_row}:H{row - 1}"

    for col_idx, width in enumerate([30, 16, 12, 14, 12, 12, 12, 18], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ---- Sheet 3+: one per entity, MID-level detail ----
    _write_entity_mid_sheets(
        wb, mid_rows, [e["entity"] for e in entities], reconciling=bool(reconciliation)
    )

    wb.remove(default_sheet)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


_SHEET_NAME_INVALID = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_name(name: str, used: set) -> str:
    """openpyxl/Excel worksheet names: no \\ / * ? : [ ], max 31 chars, unique
    within the workbook. Entity names are free text (from partner mappings),
    so any of that can show up."""
    cleaned = _SHEET_NAME_INVALID.sub("_", name or "Entity").strip() or "Entity"
    cleaned = cleaned[:31]
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


# Settlement-side columns of an entity sheet, then a narrow spacer, then the
# traced transaction. Columns A..E / G.. -- F is the gutter.
_MID_HEADERS = [
    "MID", "Merchant Name", "Acquirer Name", "Bank/Wallet Name", "Settlement Type",
    "Settled Amount", "Service Charge",
]
_MID_WIDTHS = [16, 28, 22, 22, 16, 16, 14]
# Column indices (1-based) of the settlement-side money columns, so the sheet
# can right-align and number-format them without hunting by header text.
_MID_AMOUNT_COL = _MID_HEADERS.index("Settled Amount") + 1
_MID_CHARGE_COL = _MID_HEADERS.index("Service Charge") + 1
_GUTTER_WIDTH = 3
_TXN_WIDTHS = [22, 24, 18, 14, 14, 14, 18, 22, 16, 18, 14]


def _as_number(text):
    """Amount cells arrive as text from the Transaction List ("380.00").
    Write them as numbers so the column totals and sorts like one; anything
    unparseable is written verbatim rather than dropped."""
    if text in (None, ""):
        return None
    try:
        return float(str(text).replace(",", ""))
    except ValueError:
        return text


def _write_entity_mid_sheets(
    wb, mid_rows: list[dict], entity_order: list[str], reconciling: bool = False
) -> None:
    """
    One sheet per entity (aggregator/bank/wallet/SCT/unmapped), listing every
    MID that settled under it: MID, Merchant Name, Acquirer Name and Bank/
    Wallet Name from the source file (blank where the file didn't have that
    column), and Settlement Type. Settlement Type is constrained to exactly
    the three real values via an in-cell dropdown -- a blank/unrecognized
    "Settled By" is left blank rather than guessed, for someone to fill in
    by hand before sharing.

    When a Transaction List was uploaded alongside the settlement file, each
    row also carries the transaction it was traced to, in a second table
    aligned row-for-row across the gutter column -- so the sheet reads "this
    settlement, for that transaction". A settlement with no trace still gets
    its transaction row, bordered and empty, so the two tables never drift
    out of alignment (see services/transaction_reconcile.py).

    Report-only: neither the date-range Settlement Type page nor Document
    Analysis shows MID-level detail, by request.
    """
    if not mid_rows:
        return

    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=10)
    font_title = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    font_band = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_band_settle = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    # The transaction side gets its own (teal) band and header fill so the
    # eye can tell at a glance which half of a row it is reading.
    fill_band_txn = PatternFill(start_color="1B6E62", end_color="1B6E62", fill_type="solid")
    fill_header_txn = PatternFill(start_color="2E8B7A", end_color="2E8B7A", fill_type="solid")
    fill_untraced = PatternFill(start_color="FBF3F3", end_color="FBF3F3", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    txn_first_col = len(_MID_HEADERS) + 2  # +1 for the gutter, +1 to land past it
    txn_last_col = txn_first_col + len(TXN_REPORT_HEADERS) - 1
    last_col = txn_last_col if reconciling else len(_MID_HEADERS)
    amount_idx = TXN_REPORT_KEYS.index("txn_amount")

    by_entity: dict[str, list[dict]] = defaultdict(list)
    for r in mid_rows:
        by_entity[r["entity"]].append(r)

    # Entities missing from entity_order (shouldn't happen -- both callers
    # derive mid_rows and entities from the same classification pass) still
    # get a sheet rather than being silently dropped.
    ordered = [e for e in entity_order if e in by_entity] + [
        e for e in by_entity if e not in entity_order
    ]

    used_names: set = set()

    for entity in ordered:
        rows = by_entity[entity]
        ws = wb.create_sheet(title=_safe_sheet_name(entity, used_names))

        entity_amount = sum(float(r.get("amount") or 0.0) for r in rows)
        ws["A1"] = (
            f"{entity} \u2014 Settled MIDs  "
            f"({len(rows):,} settlements, NPR {entity_amount:,.2f})"
        )
        ws["A1"].font = font_title

        header_row = 3

        if reconciling:
            # Band row directly above the headers, naming each half.
            band_row = header_row - 1
            ws.merge_cells(
                start_row=band_row, start_column=1,
                end_row=band_row, end_column=len(_MID_HEADERS),
            )
            band_a = ws.cell(row=band_row, column=1, value="Settlement File")
            band_a.font = font_band
            band_a.fill = fill_band_settle
            band_a.alignment = align_center

            ws.merge_cells(
                start_row=band_row, start_column=txn_first_col,
                end_row=band_row, end_column=txn_last_col,
            )
            band_b = ws.cell(row=band_row, column=txn_first_col, value="Traced Transaction")
            band_b.font = font_band
            band_b.fill = fill_band_txn
            band_b.alignment = align_center

        for col_idx, h in enumerate(_MID_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        if reconciling:
            for offset, h in enumerate(TXN_REPORT_HEADERS):
                cell = ws.cell(row=header_row, column=txn_first_col + offset, value=h)
                cell.font = font_header
                cell.fill = fill_header_txn
                cell.alignment = align_center

        row_idx = header_row + 1
        for r in rows:
            values = [
                r["mid"], r["merchant_name"], r["acquirer_name"],
                r["bank_wallet_name"], r["settlement_type"],
                r.get("amount", 0.0), r.get("service_charge", 0.0),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font_regular
                cell.border = border_cell
                if col_idx in (_MID_AMOUNT_COL, _MID_CHARGE_COL):
                    cell.alignment = align_right
                    cell.number_format = "#,##0.00"
                else:
                    cell.alignment = align_left

            if reconciling:
                txn = r.get("txn")
                for offset, key in enumerate(TXN_REPORT_KEYS):
                    # An untraced settlement still writes its full transaction
                    # row -- empty, bordered, tinted -- so row N on the left is
                    # always row N on the right.
                    raw = txn.get(key, "") if txn else ""
                    value = _as_number(raw) if offset == amount_idx else (raw or None)
                    cell = ws.cell(row=row_idx, column=txn_first_col + offset, value=value)
                    cell.font = font_regular
                    cell.border = border_cell
                    cell.alignment = align_right if offset == amount_idx else align_left
                    if offset == amount_idx and isinstance(value, float):
                        cell.number_format = "#,##0.00"
                    if txn is None:
                        cell.fill = fill_untraced

            row_idx += 1
        last_row = row_idx - 1

        if last_row >= header_row + 1:
            # showDropDown=False is not a typo -- in the underlying Excel XML
            # schema that flag means "suppress the dropdown arrow", so False
            # is what actually shows the in-cell picker.
            dv = DataValidation(
                type="list",
                formula1='"Real Time,System Default,On Call"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error = "Choose Real Time, System Default, or On Call."
            dv.errorTitle = "Invalid Settlement Type"
            ws.add_data_validation(dv)
            type_col = get_column_letter(_MID_HEADERS.index("Settlement Type") + 1)
            dv.add(f"{type_col}{header_row + 1}:{type_col}{last_row}")

            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

        widths = list(_MID_WIDTHS)
        if reconciling:
            widths += [_GUTTER_WIDTH] + _TXN_WIDTHS
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_settlement_type_report_bytes(date_from, date_to) -> bytes:
    """Downloadable Excel for the date-range Settlement Type report: Summary
    + Entity Breakdown + one MID-detail sheet per entity.

    No transaction pairing here: that belongs to Document Analysis, where the
    user hands over both files at once. This report covers batches already in
    the pipeline over an arbitrary date range, so there is no single
    Transaction List that would line up with it.
    """
    data = build_settlement_type_report(date_from, date_to)
    mid_rows = build_settlement_type_mid_rows(date_from, date_to)
    meta = [
        ("Date Range:", f"{data['range']['from']} to {data['range']['to']}"),
        ("Batches Included:", data["kpis"]["batches_included"]),
        ("Total Settled:", data["kpis"]["total_settled"]),
        ("Total Amount Settled:", data["kpis"]["total_amount_settled"]),
    ]
    return _write_settlement_type_workbook(meta, data, mid_rows)


def generate_adhoc_settlement_type_report_bytes(
    file_path: str, file_name: str, txn_file_path: str | None = None, txn_file_name: str = ""
) -> bytes:
    """Downloadable Excel for a standalone Document Analysis file: Summary +
    Entity Breakdown + one MID-detail sheet per entity. txn_file_path is the
    optional Transaction List uploaded beside the settlement file; when given,
    every settled MID is traced to the transaction it settled."""
    full = build_adhoc_settlement_type_full(file_path, file_name)
    data = full["summary"]
    mid_rows = full["mid_rows"]
    index = build_transaction_index(txn_file_path, txn_file_name) if txn_file_path else None
    reconciliation = attach_transactions(mid_rows, index)
    meta = [
        ("Source File:", data["file_name"]),
        ("Rows Read:", data["row_count"]),
        ("Total Settled:", data["kpis"]["total_settled"]),
        ("Total Amount Settled:", data["kpis"]["total_amount_settled"]),
    ]
    return _write_settlement_type_workbook(meta, data, mid_rows, reconciliation)


ERROR_CLASSIFY_HEADERS = [
    "Entity", "Entity Type", "Error Side", "Issue Category",
    "Failed", "Pending", "Lo Progress", "Total Txns",
    "% of Entity", "% of Batch",
]


def write_error_classify_sheet(ws, data: dict) -> None:
    """
    Writes the error-frequency table onto `ws`: one row per
    (entity, issue category), worst entity first, no charts -- the numbers
    only. Shared by the standalone extract and the report's last sheet.
    """
    batch = data["batch"]
    totals = data["totals"]

    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_entity_total = PatternFill(start_color="EDF2FA", end_color="EDF2FA", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    ws["A1"] = "Error Classification — Failure Frequency by Aggregator / SCT"
    ws["A1"].font = font_title

    ws["A2"] = (
        "Every non-success transaction in this batch, grouped by who it belongs to "
        "and what went wrong. Counts are independent of ops status (excluded and "
        "solved issues still count — they still happened). Failures that a later "
        "reprocess actually settled are excluded and counted separately below."
    )
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="808080")

    meta = [
        ("Batch:", batch["name"]),
        ("Total Error Txns:", totals["total_errors"]),
        ("Entities Affected:", totals["entities"]),
        ("Distinct Categories:", totals["categories"]),
        ("Failed / Pending / Lo Progress:", "{failed} / {pending} / {lo_progress}".format(
            **totals["by_status"]
        )),
        ("Excluded (reprocessed & settled):", totals.get("retry_resolved", 0)),
    ]
    row = 4
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=2, value=value).font = font_regular
        row += 1

    row += 1
    for col_idx, header in enumerate(ERROR_CLASSIFY_HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    header_row = row
    row += 1

    for entity in data["entities"]:
        for cat in entity["categories"]:
            values = [
                entity["entity"],
                ENTITY_TYPE_LABELS.get(entity["entity_type"], entity["entity_type"]),
                SIDE_LABELS.get(cat["side"], cat["side"]),
                cat["category"],
                cat["failed"],
                cat["pending"],
                cat["lo_progress"],
                cat["count"],
                cat["share_of_entity"] / 100.0,
                cat["share_of_batch"] / 100.0,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = font_regular
                cell.border = border_cell
                if col_idx <= 4:
                    cell.alignment = align_left
                elif col_idx <= 8:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_right
                    cell.number_format = "0.0%"
            row += 1

        # Subtotal line so a reader can scan entity-level frequency without
        # re-adding the category rows themselves.
        subtotal = [
            entity["entity"], "", "", "TOTAL",
            entity["failed"], entity["pending"], entity["lo_progress"],
            entity["total"], 1.0, entity["share_of_batch"] / 100.0,
        ]
        for col_idx, value in enumerate(subtotal, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = font_bold
            cell.fill = fill_entity_total
            cell.border = border_cell
            cell.alignment = align_left if col_idx <= 4 else align_right
            if col_idx >= 9:
                cell.number_format = "0.0%"
        row += 1

    # Entity name is repeated on every row rather than merged down the
    # column: this sheet is meant to be filtered and pivoted, and merged
    # cells break both.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:J{max(row - 1, header_row)}"

    widths = [30, 14, 12, 46, 9, 10, 12, 11, 11, 11]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _chart_row_span(height_cm: float) -> int:
    """
    How many worksheet rows a chart of this height covers, plus a gutter.

    Excel's default row height is 15pt and 1cm is 28.35pt, so a row is about
    0.53cm -- roughly 1.9 rows per cm. Charts were previously anchored to the
    row of the little source table each one reads from, but a 9cm chart spans
    ~17 rows while those tables sit 3-12 rows apart, so every chart landed on
    top of the one above it. Charts are now stacked on their own rhythm,
    independent of where their source data happens to sit.
    """
    return int(height_cm * 1.9) + 3


# Every chart is anchored in this column, so they line up in one clean strip
# down the sheet and never collide with the source tables in A-E.
_CHART_COL = "G"
_CHART_WIDTH = 17.0


def write_error_classify_charts_sheet(ws, data: dict) -> None:
    """
    Visual version of the Error Classify data for the batch report: the
    Failed/Pending/Lo Progress split, whose side the errors fall on, a
    stacked bar + cumulative-% line (Pareto) for the worst entities, and the
    most frequent issue categories.

    Layout is two columns of the sheet: narrow source tables in A-E (kept
    only because openpyxl charts must reference cells), and the charts
    themselves stacked down column G with fixed vertical spacing -- see
    _chart_row_span.

    For the full filterable per-category table, use the standalone Error
    Classify extract (write_error_classify_sheet), which this deliberately
    does not replace.
    """
    batch = data["batch"]
    totals = data["totals"]
    entities = data["entities"]
    categories = data["categories"]

    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_section = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    font_muted = Font(name="Calibri", size=9, italic=True, color="808080")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_kpi = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # A chart canvas reads better without the grid behind it.
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Error Classification — Visual Summary"
    ws["A1"].font = font_title
    ws["A2"] = (
        "Every non-success transaction in this batch, grouped by who it belongs to "
        "and what went wrong. Counts are independent of ops status -- excluded and "
        "solved issues still count, they still happened. For the full filterable "
        "per-category table, use the standalone Error Classify download."
    )
    ws["A2"].font = font_muted

    # ---- KPI strip ----
    kpis = [
        ("Batch", batch["name"]),
        ("Total Error Txns", totals["total_errors"]),
        ("Entities Affected", totals["entities"]),
        ("Distinct Categories", totals["categories"]),
        ("Reprocessed & Settled", totals.get("retry_resolved", 0)),
    ]
    row = 4
    for label, value in kpis:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = font_bold
        lc.fill = fill_kpi
        lc.border = border_cell
        lc.alignment = align_left
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = font_regular
        vc.fill = fill_kpi
        vc.border = border_cell
        vc.alignment = align_right
        row += 1

    if totals["total_errors"] == 0:
        ws.cell(row=row + 1, column=1, value="No non-success transactions in this batch.").font = font_bold
        return

    total_errors = totals["total_errors"]

    # Charts march down column G on their own rhythm; tables march down A-E on
    # theirs. The two are deliberately independent.
    chart_row = 4

    def place(chart, height_cm):
        """Anchor a chart in the chart column and reserve its vertical space."""
        nonlocal chart_row
        chart.height = height_cm
        chart.width = _CHART_WIDTH
        chart.style = 10
        ws.add_chart(chart, f"{_CHART_COL}{chart_row}")
        chart_row += _chart_row_span(height_cm)

    def write_table(title, headers, rows_data, formats=None):
        """Small source table in A-E. Returns (header_row, first_row, last_row)."""
        nonlocal row
        row += 2
        ws.cell(row=row, column=1, value=title).font = font_section
        ws.cell(row=row, column=len(headers), value="chart source").font = font_muted
        row += 1
        header_row = row
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
        row += 1
        first_row = row
        for values in rows_data:
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = font_regular
                cell.border = border_cell
                cell.alignment = align_left if col_idx == 1 else align_right
                if formats and col_idx in formats:
                    cell.number_format = formats[col_idx]
            row += 1
        return header_row, first_row, row - 1

    # ---- 1. Status distribution -> doughnut ----
    status_rows = [
        (label, totals["by_status"].get(bucket, 0))
        for bucket, label in (("failed", "Failed"), ("pending", "Pending"), ("lo_progress", "Lo Progress"))
    ]
    s_hdr, s_first, s_last = write_table("Status Distribution", ["Status", "Count"], status_rows)

    pie = PieChart()
    pie.title = f"Status Split — {total_errors:,} error transactions"
    pie.add_data(Reference(ws, min_col=2, min_row=s_hdr, max_row=s_last), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=s_first, max_row=s_last))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    place(pie, 8.0)

    # ---- 2. Which side the errors fall on -> pie ----
    side_rows = [
        (SIDE_LABELS.get(side, side), count)
        for side, count in sorted(totals.get("by_side", {}).items(), key=lambda kv: -kv[1])
        if count
    ]
    if side_rows:
        sd_hdr, sd_first, sd_last = write_table("Errors by Side", ["Side", "Count"], side_rows)
        side_pie = PieChart()
        side_pie.title = "Whose Side — SCT vs Aggregator vs Bank"
        side_pie.add_data(Reference(ws, min_col=2, min_row=sd_hdr, max_row=sd_last), titles_from_data=True)
        side_pie.set_categories(Reference(ws, min_col=1, min_row=sd_first, max_row=sd_last))
        side_pie.dataLabels = DataLabelList()
        side_pie.dataLabels.showPercent = True
        place(side_pie, 8.0)

    # ---- 3. Top entities -> stacked bar + cumulative % line (Pareto) ----
    top_entities = entities[:10]
    entity_rows = []
    running_total = 0
    for ent in top_entities:
        running_total += ent["total"]
        entity_rows.append((
            ent["entity"], ent["failed"], ent["pending"], ent["lo_progress"],
            running_total / total_errors,
        ))
    if entity_rows:
        e_hdr, e_first, e_last = write_table(
            "Top Entities by Error Volume",
            ["Entity", "Failed", "Pending", "Lo Progress", "Cumulative %"],
            entity_rows,
            formats={5: "0%"},
        )
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.overlap = 100
        bar.title = f"Top {len(top_entities)} Entities by Error Volume (Pareto)"
        bar.y_axis.title = "Transactions"
        bar.x_axis.title = "Entity"
        cats = Reference(ws, min_col=1, min_row=e_first, max_row=e_last)
        for col in (2, 3, 4):
            bar.add_data(Reference(ws, min_col=col, min_row=e_hdr, max_row=e_last), titles_from_data=True)
        bar.set_categories(cats)

        line = LineChart()
        line.add_data(Reference(ws, min_col=5, min_row=e_hdr, max_row=e_last), titles_from_data=True)
        line.set_categories(cats)
        line.y_axis.axId = 200
        line.y_axis.title = "Cumulative %"
        line.y_axis.numFmt = "0%"
        bar.y_axis.crosses = "max"
        bar += line
        place(bar, 10.0)

    # ---- 4. Top issue categories -> horizontal bar ----
    top_categories = categories[:10]
    category_rows = [
        (cat["category"], SIDE_LABELS.get(cat["side"], cat["side"]), cat["count"])
        for cat in top_categories
    ]
    if category_rows:
        c_hdr, c_first, c_last = write_table(
            "Top Issue Categories", ["Category", "Side", "Count"], category_rows
        )
        cat_bar = BarChart()
        cat_bar.type = "bar"  # horizontal: category names are long
        cat_bar.title = f"Top {len(top_categories)} Issue Categories"
        cat_bar.y_axis.title = "Transactions"
        cat_bar.x_axis.title = "Category"
        cat_bar.add_data(
            Reference(ws, min_col=3, min_row=c_hdr, max_row=c_last), titles_from_data=True
        )
        cat_bar.set_categories(Reference(ws, min_col=1, min_row=c_first, max_row=c_last))
        cat_bar.dataLabels = DataLabelList()
        cat_bar.dataLabels.showVal = True
        cat_bar.legend = None  # single series -- the title already says what it is
        place(cat_bar, 10.0)

    for col_idx, width in enumerate([34, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["F"].width = 3  # gutter between tables and charts


def generate_error_classification_bytes(batch_id: int) -> bytes:
    """Standalone one-sheet Excel of the error classification table."""
    data = build_error_classification(batch_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error Classify"
    write_error_classify_sheet(ws, data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_aggregator_report_bytes(batch_id: int, partner_name: str, status_filter: str = None) -> bytes:
    """
    Generates a targeted extract for a specific aggregator containing ONLY
    transactions whose resolved issue status is pending, failed, or lo_progress.
    Optionally filters by the original transaction status.
    """
    batch = Batch.query.get_or_404(batch_id)
    transactions = Transaction.query.filter_by(batch_id=batch_id, partner_name=partner_name).all()
    issues = IssueStatus.query.filter_by(batch_id=batch_id).all()
    
    issue_status_map = {
        (i.side, i.partner_name, i.category, i.txn_status): i
        for i in issues
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aggregator Extract"
    
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=10)
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    orig_headers = []
    if batch.input_file_path and os.path.exists(batch.input_file_path):
        try:
            h_idx = _find_header_index(batch.input_file_path)
            raw_df = pd.read_excel(batch.input_file_path, header=h_idx, nrows=1)
            orig_headers = [str(c).strip() for c in raw_df.columns if pd.notna(c)]
        except Exception:
            pass

    if not orig_headers:
        orig_headers = ["SN", "Transaction Date", "Acquirer Name", "MID", "Merchant Name", "Txn Amount", "Service Charge", "Settled By", "STAN", "CRRN", "CR Transaction ID", "Bank/Wallet Name", "Bank Account/Wallet ID", "Account Name", "Remarks 1", "Remarks 2", "Status Code", "Status"]

    appended_headers = ["Issue Category", "Solved Status", "Ops Remarks"]
    full_headers = orig_headers + appended_headers

    for col_idx, h in enumerate(full_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    row_idx = 2
    for txn in transactions:
        if txn.retry_resolved:
            continue  # already settled on a retry -- nothing to chase up

        txn_original_status = normalize_txn_status(txn.status)
        if status_filter and txn_original_status != status_filter:
            continue

        extra = txn.extra_data or {}
        issue_obj = issue_status_map.get((txn.error_side, txn.partner_name if txn.error_side != "sct" else None, txn.error_category, txn_original_status))
        
        override_obj = None
        if issue_obj and issue_obj.mid_overrides and txn.mid in issue_obj.mid_overrides:
            raw = issue_obj.mid_overrides[txn.mid]
            if isinstance(raw, dict):
                override_obj = raw
            else:
                override_obj = {"status": raw, "remark": ""}

        eff_status = (override_obj["status"] if override_obj else None) or (issue_obj.status if issue_obj else "pending")
        
        # Only include un-solved, non-excluded transactions
        if eff_status in ("solved", "exclude", "success"):
            continue

        solved_status = eff_status.replace("_", " ").title()
        if override_obj and override_obj.get("remark"):
            solved_remarks = f"[MID Override] {override_obj['remark']}"
        else:
            solved_remarks = issue_obj.comment if issue_obj else ""

        row_vals = []
        for h in orig_headers:
            if h == "MID": row_vals.append(txn.mid)
            elif h == "Merchant Name": row_vals.append(txn.merchant_name)
            elif h == "Remarks 1": row_vals.append(txn.remark)
            elif h == "Status Code": row_vals.append(txn.status_code)
            elif h == "Status": row_vals.append(txn.status)
            else: row_vals.append(extra.get(h, ""))
            
        row_vals.extend([
            txn.error_category or "Unclassified",
            solved_status,
            solved_remarks
        ])

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            cell.alignment = align_left

        row_idx += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if val_str:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==========================================
# Issuer / Acquirer reconciliation report
# ==========================================

def generate_issuer_acquirer_report_bytes(data: dict, reasons: dict | None = None) -> bytes:
    """
    Three sheets: Summary, Issuing, Acquiring.

    Issuing and Acquiring are separate sheets by request -- they answer
    different questions and are read by different people. Issuing is "whose
    customers spent", one row per issuing bank or wallet. Acquiring is "who got
    paid", and carries the reconciliation: what the switch transacted, what
    actually settled, and the gap between them.

    `reasons` maps an acquirer name to the note ops typed against its variance.
    A number cannot say whether a shortfall is timing or a break, so the column
    exists to be filled in before the report is shared.
    """
    reasons = reasons or {}
    totals = data["totals"]

    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_section = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    font_muted = Font(name="Calibri", size=9, italic=True, color="808080")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    MONEY = "#,##0.00"

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    def header_row(ws, row, headers, widths):
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def write(ws, row, values, formats=None):
        for idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=idx, value=v)
            cell.font = font_regular
            cell.border = border_cell
            cell.alignment = align_left if idx == 1 else align_right
            if formats and idx in formats:
                cell.number_format = formats[idx]

    # ---- Sheet 1: Summary ----
    ws = wb.create_sheet(title="Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Issuer / Acquirer Reconciliation"
    ws["A1"].font = font_title
    ws["A2"] = (
        "Every transaction debits an issuer and pays an acquirer, so inside the "
        "transaction file those two sides always balance. The gap that matters is "
        "against settlement: what was transacted versus what was actually paid out."
    )
    ws["A2"].font = font_muted

    row = 4
    for label, value, fmt in (
        ("Period covered:", totals.get("window") or "—", None),
        ("Transactions read:", totals["txn_rows"], "#,##0"),
        ("Transaction amount:", totals["txn_amount"], MONEY),
        ("Settlement rows read:", totals["settlement_rows"], "#,##0"),
        ("Settled successfully:", totals["settled_count"], "#,##0"),
        ("Settled amount:", totals["settled_amount"], MONEY),
        ("Variance (transacted - settled):", totals["variance_amount"], MONEY),
        ("Settled share of transacted:", totals["settled_pct"] / 100.0, "0.0%"),
    ):
        ws.cell(row=row, column=1, value=label).font = font_bold
        c = ws.cell(row=row, column=2, value=value)
        c.font = font_regular
        c.alignment = align_right
        if fmt:
            c.number_format = fmt
        row += 1

    if not totals.get("has_settlement"):
        row += 1
        n = ws.cell(row=row, column=1, value=(
            "No settlement file was uploaded, so the settled columns and every "
            "variance are blank. The issuing and acquiring splits below still hold."
        ))
        n.font = font_muted
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    # ---- Sheet 2: Issuing ----
    ws = wb.create_sheet(title="Issuing")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Issuing — by issuing bank / wallet"
    ws["A1"].font = font_title
    ws["A2"] = "Whose customers spent. One row per issuer named by the switch."
    ws["A2"].font = font_muted
    header_row(ws, 4, ["Issuer", "Transactions", "Amount", "% of total"], [40, 15, 20, 13])
    row = 5
    for r in data["issuing"]:
        write(ws, row, [r["name"], r["txn_count"], r["txn_amount"], r["share"] / 100.0],
              formats={2: "#,##0", 3: MONEY, 4: "0.0%"})
        row += 1
    if data["issuing"]:
        ws.cell(row=row, column=1, value="Total").font = font_bold
        for col, val, fmt in ((2, totals["txn_rows"], "#,##0"), (3, totals["txn_amount"], MONEY)):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.number_format = fmt
            c.alignment = align_right
        ws.freeze_panes = ws.cell(row=5, column=1)
        ws.auto_filter.ref = f"A4:D{row - 1}"

    # ---- Sheet 3: Acquiring ----
    ws = wb.create_sheet(title="Acquiring")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Acquiring — transacted vs settled"
    ws["A1"].font = font_title
    ws["A2"] = (
        "Who got paid. A positive variance is transacted-but-not-yet-settled; "
        "negative means the settlement file covers transactions this export does not. "
        "Fill in the reason before sharing — the number alone cannot say whether a gap "
        "is timing or a break."
    )
    ws["A2"].font = font_muted
    header_row(
        ws, 4,
        ["Acquirer", "Txns", "Transacted", "Settled txns", "Settled", "Variance", "Reason"],
        [36, 10, 18, 13, 18, 18, 44],
    )
    row = 5
    for r in data["acquiring"]:
        write(ws, row, [
            r["name"], r["txn_count"], r["txn_amount"],
            r["settled_count"], r["settled_amount"], r["variance_amount"],
            reasons.get(r["name"], ""),
        ], formats={2: "#,##0", 3: MONEY, 4: "#,##0", 5: MONEY, 6: MONEY})
        ws.cell(row=row, column=7).alignment = align_left
        row += 1
    if data["acquiring"]:
        ws.cell(row=row, column=1, value="Total").font = font_bold
        for col, val in ((2, totals["txn_rows"]), (3, totals["txn_amount"]),
                         (4, totals["settled_count"]), (5, totals["settled_amount"]),
                         (6, totals["variance_amount"])):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.number_format = MONEY if col in (3, 5, 6) else "#,##0"
            c.alignment = align_right
        ws.freeze_panes = ws.cell(row=5, column=1)
        ws.auto_filter.ref = f"A4:G{row - 1}"

    wb.remove(default_sheet)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
